import openpyxl

path = "C:\\Users\\pc\\Desktop\\DDF_V1.xlsx"

wb = openpyxl.load_workbook(path)
sh = wb['Sheet5']
print("====================")
# row count and column count max_row, max_column
print("row count = ", sh.max_row)
print("Column count = ",sh.max_column)
print("====================")
# read the data --> cell(row= , column=).value
print(sh.cell(row=4,column=1).value)
print(sh.cell(row=6, column=2).value)

# print only one column
print("======== username ============")
for i in range(1, sh.max_row+1):
    print(sh.cell(row=i, column=1).value)

print("========= password ===========")
for i in range(1, sh.max_row+1):
    print(sh.cell(row=i, column=2).value)

print("=======all the rows and columns=========")

for i in range(1,sh.max_row+1):
    for j in range(1, sh.max_column+1):
        print(sh.cell(row=i, column=j).value)