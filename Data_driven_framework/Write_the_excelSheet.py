import openpyxl
path = "C:\\Users\\pc\\Desktop\\DDF_V1.xlsx"
wb = openpyxl.load_workbook(path)
sh = wb['Sheet6']
sh.cell(row=1,column=1).value="Python"

for i in range(2, 7):
    sh.cell(row=i, column=1).value="java"

for i in range(7, 13):
    for j in range(1, 4):
        sh.cell(row=i,column=j).value="Karthik"

wb.save(path)
print("====written successfully====")