import openpyxl


def getRowCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheetName]
    return sh.max_row


def getColumnCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheetName]
    return sh.max_column


def readData(file, sheetName, r, c):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheetName]
    return sh.cell(row=r, column=c).value


def writeData(file, sheetName, r, c, val):
    wb = openpyxl.load_workbook(file)
    sh = wb[sheetName]
    sh.cell(row=r, column=c).value = val
    wb.save(file)
